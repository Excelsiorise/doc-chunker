from __future__ import annotations

import csv
import zipfile
from pathlib import Path
from typing import Any
from xml.etree import ElementTree as ET

from doc_chunker.models import DocumentBlock

WORD_NS = "{http://schemas.openxmlformats.org/wordprocessingml/2006/main}"


def parse_document(path: str | Path) -> list[DocumentBlock]:
    file_path = Path(path)
    suffix = file_path.suffix.lower()
    if suffix in {".txt", ".md"}:
        return parse_text(file_path)
    if suffix == ".docx":
        return parse_docx(file_path)
    if suffix == ".xlsx":
        return parse_xlsx(file_path)
    if suffix == ".csv":
        return parse_csv(file_path)
    if suffix == ".pdf":
        return parse_pdf(file_path)
    raise ValueError(f"Unsupported document type: {suffix}")


def parser_name(path: str | Path) -> str:
    return Path(path).suffix.lower().lstrip(".") or "text"


def parse_text(path: Path) -> list[DocumentBlock]:
    text = path.read_text(encoding="utf-8")
    blocks: list[DocumentBlock] = []
    for i, para in enumerate([p.strip() for p in text.split("\n\n") if p.strip()], start=1):
        blocks.append(
            DocumentBlock(
                text=para,
                source_file=str(path),
                block_type="paragraph",
                locator={"paragraph": i},
            )
        )
    return blocks


def parse_docx(path: Path) -> list[DocumentBlock]:
    with zipfile.ZipFile(path) as zf:
        xml = zf.read("word/document.xml")
    root = ET.fromstring(xml)
    blocks: list[DocumentBlock] = []
    heading_path: list[str] = []
    paragraph_no = 0
    for para in root.iter(f"{WORD_NS}p"):
        paragraph_no += 1
        text = "".join(node.text or "" for node in para.iter(f"{WORD_NS}t")).strip()
        if not text:
            continue
        style = _paragraph_style(para)
        level = _heading_level(style)
        if level:
            heading_path = heading_path[: level - 1] + [text]
            block_type = "heading"
            block_heading = heading_path[: level - 1]
        else:
            block_type = "paragraph"
            block_heading = list(heading_path)
        blocks.append(
            DocumentBlock(
                text=text,
                source_file=str(path),
                block_type=block_type,
                locator={"paragraph": paragraph_no},
                heading_path=block_heading,
                metadata={"style": style} if style else {},
            )
        )
    return blocks


def parse_xlsx(path: Path) -> list[DocumentBlock]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - dependency declared in pyproject
        raise RuntimeError("openpyxl is required to parse .xlsx files") from exc

    wb = load_workbook(path, read_only=True, data_only=True)
    blocks: list[DocumentBlock] = []
    try:
        for ws in wb.worksheets:
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue
            headers = [_cell_to_text(value) for value in rows[0]]
            for row_index, row in enumerate(rows[1:], start=2):
                values = [_cell_to_text(value) for value in row]
                if not any(values):
                    continue
                pairs = []
                for header, value in zip(headers, values):
                    label = header or "Column"
                    if value:
                        pairs.append(f"{label}: {value}")
                blocks.append(
                    DocumentBlock(
                        text="; ".join(pairs) if pairs else "; ".join(values),
                        source_file=str(path),
                        block_type="table_row",
                        locator={"sheet": ws.title, "row": row_index},
                        heading_path=[ws.title],
                        metadata={"headers": headers},
                    )
                )
    finally:
        wb.close()
    return blocks


def parse_csv(path: Path) -> list[DocumentBlock]:
    blocks: list[DocumentBlock] = []
    with path.open("r", encoding="utf-8", newline="") as handle:
        reader = csv.reader(handle)
        rows = list(reader)
    if not rows:
        return blocks
    headers = rows[0]
    for row_index, row in enumerate(rows[1:], start=2):
        text = "; ".join(f"{header}: {value}" for header, value in zip(headers, row) if value)
        if text:
            blocks.append(
                DocumentBlock(
                    text=text,
                    source_file=str(path),
                    block_type="table_row",
                    locator={"row": row_index},
                    metadata={"headers": headers},
                )
            )
    return blocks


def parse_pdf(path: Path) -> list[DocumentBlock]:
    try:
        from pypdf import PdfReader
    except ImportError as exc:  # pragma: no cover - dependency declared in pyproject
        raise RuntimeError("pypdf is required to parse .pdf files") from exc

    reader = PdfReader(str(path))
    blocks: list[DocumentBlock] = []
    for page_index, page in enumerate(reader.pages, start=1):
        text = (page.extract_text() or "").strip()
        if text:
            blocks.append(
                DocumentBlock(
                    text=text,
                    source_file=str(path),
                    block_type="page",
                    locator={"page": page_index},
                )
            )
    return blocks


def _paragraph_style(para: ET.Element) -> str:
    ppr = para.find(f"{WORD_NS}pPr")
    if ppr is None:
        return ""
    style = ppr.find(f"{WORD_NS}pStyle")
    if style is None:
        return ""
    return style.attrib.get(f"{WORD_NS}val", "")


def _heading_level(style: str) -> int | None:
    lower = style.lower()
    if lower.startswith("heading"):
        suffix = lower.removeprefix("heading")
        if suffix.isdigit():
            return max(1, int(suffix))
        return 1
    return None


def _cell_to_text(value: Any) -> str:
    return "" if value is None else str(value).strip()
